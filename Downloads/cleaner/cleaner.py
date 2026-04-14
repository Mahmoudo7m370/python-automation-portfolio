import os
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font
if not os.path.exists("messy_data.xlsx"):
    print("Error: messy_data.xlsx not found.")
    exit()
wb=load_workbook("messy_data.xlsx")
ws=wb.active
rows=[]
empty_cells=0
header=[]
for cell in ws[1]:
    header.append(cell.value)

for row in ws.iter_rows(min_row=2,values_only=True):
    cleaned=[]
    for cell in row:
        if cell is None:
            cell="N/A"
            empty_cells+=1
        elif isinstance(cell,str):
           cell= cell.strip().title()
        cleaned.append(cell)
    rows.append(tuple(cleaned))
unique=[]
seen=set()
for row in rows:
  if row not in seen:
      seen.add(row)
      unique.append(row)  
duplicates=len(rows)-len(unique)
wb2=Workbook()
ws2=wb2.active
ws2.append(header)
for cell in ws2[1]:
    cell.font=Font(bold=True)
for row in unique:
    ws2.append(list(row))
print(f"Removed {duplicates} duplicates. Fixed {empty_cells} empty cells.")
wb2.save("cleaned_data.xlsx")