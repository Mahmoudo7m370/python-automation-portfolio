from openpyxl import load_workbook
from openpyxl import Workbook 
from openpyxl.styles import Font
import glob
files=glob.glob("*.xlsx")
clean_files=[]
for f in files:
    if f !="master_report.xlsx":
        clean_files.append(f)
files=clean_files
rows=[]
i=0
header=None
for file in files:
    wb=load_workbook(file)
    ws=wb.active
    for  row in ws.iter_rows(min_row=1,values_only=True):
        header_row=row
        break
    if "Amount"  not in header_row:
        continue
    if header is None:
        header=header_row
        i=header.index("Amount")
    for row in ws.iter_rows(min_row=2,values_only=True):
       rows.append(row) 
if header is None:
    print("No valid files found with 'Amount' column.")
    exit()
rows=sorted(rows,key=lambda x:x[i],reverse=True)
wb2=Workbook()
ws2=wb2.active
ws2.append(header)
for cell in ws2[1]:
    cell.font=Font(bold=True)
for row in rows:
    ws2.append(list(row))
print( f"Merged {len(files)} files. Total rows:{len(rows)}.")
wb2.save("master_report.xlsx")