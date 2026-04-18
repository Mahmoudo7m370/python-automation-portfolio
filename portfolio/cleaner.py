import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# Check if the input file exists before doing anything
if not os.path.exists("messy_data.xlsx"):
    print("Error: messy_data.xlsx not found.")
    exit()

# Load the messy Excel file
wb = load_workbook("messy_data.xlsx")
ws = wb.active

# Read the header row
rows = []
empty_cells = 0
header = []
for cell in ws[1]:
    header.append(cell.value)

# Clean each row — fix casing, strip spaces, fill empty cells
for row in ws.iter_rows(min_row=2, values_only=True):
    cleaned = []
    for cell in row:
        if cell is None:
            cell = "N/A"
            empty_cells += 1
        elif isinstance(cell, str):
            cell = cell.strip().title()
        cleaned.append(cell)
    rows.append(tuple(cleaned))

# Remove duplicate rows
unique = []
seen = set()
for row in rows:
    if row not in seen:
        seen.add(row)
        unique.append(row)
duplicates = len(rows) - len(unique)

# Write cleaned data to a new file with bold headers
wb2 = Workbook()
ws2 = wb2.active
ws2.append(header)
for cell in ws2[1]:
    cell.font = Font(bold=True)
for row in unique:
    ws2.append(list(row))

# Print summary and save — never overwrites the original
print(f"Removed {duplicates} duplicates. Fixed {empty_cells} empty cells.")
wb2.save("cleaned_data.xlsx")
