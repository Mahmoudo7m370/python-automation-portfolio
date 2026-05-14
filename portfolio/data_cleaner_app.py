import io
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ── Page Header ───────────────────────────────────────────────
st.title("Data Cleaner Pro")
st.write("Upload your messy CSV or Excel files and get a clean, formatted Excel report.")

# ── Sidebar: App Settings ──────────────────────────────────────
st.sidebar.header("Settings")
selected_mode = st.sidebar.selectbox("Choose mode", ["Clean Only", "Clean + Summary Report"])

# ── Main Area: File Upload ─────────────────────────────────────
uploaded_files = st.file_uploader(
    "Upload CSV or Excel files",
    type=["csv", "xlsx"],
    accept_multiple_files=True
)

if uploaded_files:

    enable_highlight = False
    progress_bar = st.progress(0)

    with st.spinner("Cleaning your data..."):

        # ── Step 1: Read and Combine All Uploaded Files ────────
        combined_frames = []
        for uploaded_file in uploaded_files:
            try:
                combined_frames.append(pd.read_csv(uploaded_file))
            except Exception:
                try:
                    combined_frames.append(pd.read_excel(uploaded_file))
                except Exception:
                    st.error(f"Could not read file: {uploaded_file.name}")
                    st.stop()

        cleaned_df = pd.concat(combined_frames, ignore_index=True)
        progress_bar.progress(20)

        # ── Step 2: Track Stats Before Cleaning ───────────────
        row_count_before = len(cleaned_df)
        empty_cell_count = int(cleaned_df.isnull().sum().sum())

        # ── Step 3: Clean Text Columns ────────────────────────
        for col in cleaned_df.columns:
            try:
                cleaned_df[col] = cleaned_df[col].str.strip().str.title()
            except Exception:
                pass

        # ── Step 4: Remove Duplicate Rows ─────────────────────
        cleaned_df = cleaned_df.drop_duplicates()
        if len(cleaned_df) == 0:
            st.error("No data remaining after cleaning.")
            st.stop()

        # ── Step 5: Validate Mode Requirements ────────────────
        has_numeric = len(cleaned_df.select_dtypes(include="number").columns) > 0
        if selected_mode == "Clean + Summary Report" and not has_numeric:
            st.error("No numeric columns found. Summary mode requires at least one numeric column.")
            st.stop()

        # ── Step 6: Fill Empty Cells ───────────────────────────
        numeric_columns = cleaned_df.select_dtypes(include="number").columns
        cleaned_df[numeric_columns] = cleaned_df[numeric_columns].fillna(0)
        cleaned_df = cleaned_df.fillna("N/A")

        progress_bar.progress(40)

        # ── Step 7: Optional Sorting ───────────────────────────
        sort_column = st.sidebar.selectbox(
            "Sort by column",
            ["None"] + list(cleaned_df.columns),
            key="main_sort_column"
        )
        if sort_column != "None":
            sort_order = st.sidebar.selectbox(
                "Sort order",
                ["Ascending", "Descending"],
                key="main_sort_direction"
            )
            try:
                cleaned_df = cleaned_df.sort_values(
                    sort_column,
                    ascending=(sort_order == "Ascending")
                )
            except Exception:
                st.warning("Could not sort — mixed data types in that column.")

        progress_bar.progress(60)

        # ── Step 8: Display Cleaned Data ──────────────────────
        st.subheader("Cleaned Data")
        st.dataframe(cleaned_df)

        duplicates_removed = row_count_before - len(cleaned_df)
        summary_df = None

        # ── Step 9: Summary Report Mode ───────────────────────
        if selected_mode == "Clean + Summary Report":

            num_cols = list(cleaned_df.select_dtypes(include="number").columns)
            cat_cols = list(cleaned_df.select_dtypes(exclude="number").columns)

            if not num_cols or not cat_cols:
                st.warning("Not enough columns for a summary report.")
            else:
                # Let user choose both columns
                groupby_column = st.sidebar.selectbox(
                    "Group by (category column)",
                    cat_cols,
                    key="groupby_col"
                )
                amount_column = st.sidebar.selectbox(
                    "Analyze (numeric column)",
                    num_cols,
                    key="summary_col"
                )

                cleaned_df[amount_column] = pd.to_numeric(
                    cleaned_df[amount_column], errors="coerce"
                )

                # Build summary
                summary_df = cleaned_df.groupby(groupby_column)[amount_column].agg(
                    Total="sum",
                    Average="mean",
                    Highest="max"
                ).reset_index()

                summary_df["Total"] = summary_df["Total"].round(2)
                summary_df["Average"] = summary_df["Average"].round(2)
                summary_df["Highest"] = summary_df["Highest"].round(2)

                # Optional sort on summary
                summary_sort = st.sidebar.selectbox(
                    "Sort summary by",
                    ["None", "Total", "Average", "Highest"],
                    key="summary_sort_col"
                )
                if summary_sort != "None":
                    summary_sort_order = st.sidebar.selectbox(
                        "Summary sort order",
                        ["Descending", "Ascending"],
                        key="summary_sort_dir"
                    )
                    summary_df = summary_df.sort_values(
                        summary_sort,
                        ascending=(summary_sort_order == "Ascending")
                    )

                st.subheader("Summary Report")
                st.dataframe(summary_df)

                # ── Bar Chart ──────────────────────────────────
                chart_metric = st.selectbox(
                    "Chart metric",
                    ["Total", "Average", "Highest"],
                    key="chart_metric"
                )
                chart_data = summary_df.set_index(groupby_column)[[chart_metric]]
                st.subheader(f"{chart_metric} by {groupby_column}")
                st.bar_chart(chart_data)

        progress_bar.progress(80)

        # ── Step 10: Optional Row Highlight ───────────────────
        enable_highlight = st.sidebar.checkbox("Highlight a row", key="highlight_toggle")
        if enable_highlight:
            highlight_row_number = st.sidebar.number_input(
                "Row to highlight (1 = first row)",
                min_value=1,
                max_value=len(cleaned_df),
                step=1,
                key="highlight_row"
            )
            highlight_hex_color = st.sidebar.color_picker(
                "Highlight color",
                "#FFFF00",
                key="highlight_color"
            )

        # ── Step 11: Write to Excel in Memory ─────────────────
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer) as excel_writer:
            cleaned_df.to_excel(excel_writer, sheet_name="Cleaned Data", index=False)
            if summary_df is not None:
                summary_df.to_excel(excel_writer, sheet_name="Summary", index=False)

        # ── Step 12: Apply Formatting with openpyxl ───────────
        excel_buffer.seek(0)
        workbook = load_workbook(excel_buffer)
        cleaned_sheet = workbook["Cleaned Data"]

        for cell in cleaned_sheet[1]:
            cell.font = Font(bold=True)

        if summary_df is not None:
            for cell in workbook["Summary"][1]:
                cell.font = Font(bold=True)

        if enable_highlight:
            row_fill_color = highlight_hex_color.lstrip("#")
            row_fill = PatternFill(
                start_color=row_fill_color,
                end_color=row_fill_color,
                fill_type="solid"
            )
            for cell in cleaned_sheet[int(highlight_row_number) + 1]:
                cell.fill = row_fill

        # ── Step 13: Save Final Workbook ──────────────────────
        final_output = io.BytesIO()
        workbook.save(final_output)
        progress_bar.progress(100)
        final_output.seek(0)

    # ── Stats + Download ──────────────────────────────────────
    col1, col2, col3 = st.columns(3)
    col1.metric("Rows Cleaned", len(cleaned_df))
    col2.metric("Duplicates Removed", duplicates_removed)
    col3.metric("Empty Cells Fixed", empty_cell_count)

    st.success("Your file is ready to download!")
    st.download_button(
        "⬇️ Download Cleaned Excel",
        final_output.getvalue(),
        "cleaned.xlsx"
    )

else:
    st.write("Please upload a file to get started.")
