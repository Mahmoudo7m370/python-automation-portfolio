from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd
import glob

# Find all Excel files in the current folder
files = glob.glob("*.xlsx")
if not files:
    print("No excel files found!")
    exit()

# Read and merge all Excel files, skipping the output file
df = pd.concat((pd.read_excel(f) for f in files if f != "final_report.xlsx"), ignore_index=True)

# Count and remove duplicate rows
duplicates = len(df) - len(df.drop_duplicates())
df = df.drop_duplicates()

# Count empty cells before cleaning
empty_cells = df.isnull().sum().sum()

# Clean text columns — strip spaces and fix casing
# For number columns — convert to numeric and drop rows with empty values
for col in df.columns:
    try:
        df[col] = df[col].str.strip().str.title()
    except:
        df[col] = pd.to_numeric(df[col], errors="coerce")
        df = df.dropna(subset=[col])

# Count empty cells after cleaning
empty_cells = df.isnull().sum().sum()

# Generate sales summary grouped by salesperson
summary = df.groupby("sales person")["sale amount"].agg(
    Total="sum", Average="mean", Highest="max"
).reset_index()
summary["Average"] = summary["Average"].round()
summary = summary.sort_values("Total", ascending=False)

# Get top salesperson name
top = summary.iloc[0]["sales person"]

# Fill any remaining empty cells
summary = summary.fillna("N/A")

# Save summary to Excel
summary.to_excel("final_report.xlsx", index=False)

# Open the saved file and apply formatting
wb = load_workbook("final_report.xlsx")
ws = wb.active

# Style definitions
green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Bold all header cells in row 1
for cell in ws[1]:
    cell.font = Font(bold=True)

# Highlight top salesperson row in green
for cell in ws[2]:
    cell.fill = green

# Set column width for readability
for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = 15

# Save final formatted report
wb.save("final_report.xlsx")
