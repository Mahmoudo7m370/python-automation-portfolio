from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import glob

# Find all Excel files in the current folder
files = glob.glob("*.xlsx")

# Remove the output file from the list to avoid merging it into itself
clean_files = []
for f in files:
    if f != "master_report.xlsx":
        clean_files.append(f)
files = clean_files

rows = []
i = 0
header = None

# Loop through each file and collect all data rows
for file in files:
    wb = load_workbook(file)
    ws = wb.active

    # Read the header row
    for row in ws.iter_rows(min_row=1, values_only=True):
        header_row = row
        break

    # Skip files that don't have an Amount column
    if "Amount" not in header_row:
        continue

    # Save the header and find the Amount column index
    if header is None:
        header = header_row
        i = header.index("Amount")

    # Collect all data rows from this file
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(row)

# Exit if no valid files were found
if header is None:
    print("No valid files found with 'Amount' column.")
    exit()

# Sort all rows by Amount — highest first
rows = sorted(rows, key=lambda x: x[i], reverse=True)

# Write merged data to master_report.xlsx with bold headers
wb2 = Workbook()
ws2 = wb2.active
ws2.append(header)
for cell in ws2[1]:
    cell.font = Font(bold=True)
for row in rows:
    ws2.append(list(row))

# Save and print summary
wb2.save("master_report.xlsx")
print(f"Merged {len(files)} files. Total rows: {len(rows)}.")
