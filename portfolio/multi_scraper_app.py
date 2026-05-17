import io
import re
import time
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import requests
from bs4 import BeautifulSoup
import plotly.express as px

st.title("Multi-Page Web Data Extractor")
st.write("Extract data from multiple pages and download a clean Excel report.")

# ── Sidebar ──────────────────────────────────────────────
st.sidebar.header("Settings")

mode = st.sidebar.selectbox("Report Mode", ["Extract Data Only", "Full Business Report"])

scraping_method = st.sidebar.radio(
    "Scraping Mode",
    ["Auto (Tables)", "Smart Extract (Books / Products)", "Table Parser (Wikipedia / Any Table)", "Advanced (CSS Selector)"]
)

total_pages = st.sidebar.number_input("Number of pages", min_value=1, max_value=50, value=3)

url_template = st.sidebar.text_input("URL pattern (use {page})", placeholder="https://example.com/page-{page}.html")

if not url_template:
    st.info("Enter a URL pattern to start.")
    st.stop()

req_headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ── Validate first page ──────────────────────────────────
first_url = url_template.replace("{page}", "1")
try:
    first_response = requests.get(first_url, headers=req_headers, timeout=15)
    html = first_response.text
except Exception:
    st.error("❌ Could not access the website.")
    st.stop()

if first_response.status_code != 200:
    st.error(f"❌ Access denied ({first_response.status_code})")
    st.stop()

st.success(f"✅ Page loaded ({len(html):,} characters)")

# ── Mode-specific setup ───────────────────────────────────
selected_table_index = 0
selector = None
table_parser_index = 0

if scraping_method == "Auto (Tables)":
    try:
        tables = pd.read_html(io.StringIO(html))
        st.success(f"✅ Found {len(tables)} table(s) on page 1")
        selected_table_index = st.sidebar.selectbox("Select Table", range(len(tables)), format_func=lambda x: f"Table {x+1} ({len(tables[x])} rows)")
    except Exception:
        st.warning("⚠️ No tables found. Try Table Parser or Smart Extract.")
        st.stop()

elif scraping_method == "Table Parser (Wikipedia / Any Table)":
    soup_first = BeautifulSoup(html, "html.parser")
    all_tables = soup_first.find_all("table")
    if not all_tables:
        st.error("❌ No tables found.")
        st.stop()
    st.info(f"Found {len(all_tables)} table(s) on page 1.")
    table_labels = []
    for i, t in enumerate(all_tables):
        first_row = t.find("tr")
        if first_row:
            cells = [c.get_text(strip=True)[:20] for c in first_row.find_all(["th", "td"])[:4]]
            table_labels.append(f"Table {i+1}: {', '.join(cells)}")
        else:
            table_labels.append(f"Table {i+1}")
    table_parser_index = st.sidebar.selectbox("Select Table", range(len(all_tables)), format_func=lambda x: table_labels[x])

elif scraping_method == "Advanced (CSS Selector)":
    selector = st.text_input("CSS Selector (e.g. div.title)")
    if not selector:
        st.info("Enter a CSS selector.")
        st.stop()

# ── Scrape all pages ──────────────────────────────────────
all_data = []
progress = st.progress(0)
status = st.empty()
rating_map = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}

for i in range(1, total_pages + 1):
    url = url_template.replace("{page}", str(i))
    status.text(f"Scraping page {i}/{total_pages}...")
    try:
        res = requests.get(url, headers=req_headers, timeout=15)
        page_html = res.text
        if res.status_code != 200:
            st.warning(f"⚠️ Page {i} returned {res.status_code}")
            continue

        if scraping_method == "Auto (Tables)":
            tables = pd.read_html(io.StringIO(page_html))
            all_data.append(tables[selected_table_index])

        elif scraping_method == "Smart Extract (Books / Products)":
            soup = BeautifulSoup(page_html, "html.parser")
            products = soup.select("article.product_pod")
            if not products:
                st.warning(f"⚠️ No products on page {i}")
                continue
            rows = []
            for p in products:
                title_tag = p.select_one("h3 a")
                title = title_tag["title"] if title_tag and title_tag.has_attr("title") else (title_tag.get_text(strip=True) if title_tag else "N/A")
                price_tag = p.select_one("p.price_color")
                price_text = price_tag.get_text(strip=True) if price_tag else "0"
                price = float(re.sub(r"[^\d.]", "", price_text)) if price_text else 0.0
                rating_tag = p.select_one("p.star-rating")
                rating_word = rating_tag["class"][1] if rating_tag and len(rating_tag["class"]) > 1 else "Zero"
                avail_tag = p.select_one("p.availability")
                rows.append({"Title": title, "Price (£)": price, "Rating": rating_map.get(rating_word, 0), "Availability": avail_tag.get_text(strip=True) if avail_tag else "N/A"})
            all_data.append(pd.DataFrame(rows))

        elif scraping_method == "Table Parser (Wikipedia / Any Table)":
            soup = BeautifulSoup(page_html, "html.parser")
            page_tables = soup.find_all("table")
            if not page_tables or table_parser_index >= len(page_tables):
                st.warning(f"⚠️ Table not found on page {i}")
                continue
            selected_table = page_tables[table_parser_index]
            rows = []
            page_headers = []
            for row in selected_table.find_all("tr"):
                cells = row.find_all(["th", "td"])
                cell_values = [c.get_text(strip=True) for c in cells]
                if not cell_values:
                    continue
                if not page_headers:
                    page_headers = cell_values
                else:
                    while len(cell_values) < len(page_headers):
                        cell_values.append("")
                    rows.append(cell_values[:len(page_headers)])
            if rows and page_headers:
                clean_headers = []
                seen = {}
                for h in page_headers:
                    h = h.strip() if h.strip() else "Column"
                    if h in seen:
                        seen[h] += 1
                        h = f"{h}_{seen[h]}"
                    else:
                        seen[h] = 0
                    clean_headers.append(h)
                df_page = pd.DataFrame(rows, columns=clean_headers)
                for col in df_page.columns:
                    try:
                        cleaned = df_page[col].str.replace(r"\[.*?\]", "", regex=True).str.replace(r"[,$€£%†♠]", "", regex=True).str.strip()
                        df_page[col] = pd.to_numeric(cleaned, errors="raise")
                    except Exception:
                        pass
                all_data.append(df_page)

        elif scraping_method == "Advanced (CSS Selector)":
            soup = BeautifulSoup(page_html, "html.parser")
            elements = soup.select(selector)
            if not elements:
                st.warning(f"⚠️ No elements on page {i}")
                continue
            all_data.append(pd.DataFrame([el.get_text(strip=True) for el in elements], columns=["Extracted Data"]))

    except Exception as e:
        st.warning(f"⚠️ Failed page {i}: {str(e)[:60]}")
        continue

    progress.progress(int((i / total_pages) * 60))
    time.sleep(0.3)

status.empty()

if not all_data:
    st.error("❌ No data scraped.")
    st.stop()

df_raw = pd.concat(all_data, ignore_index=True)

st.subheader("🔍 Raw Preview")
st.dataframe(df_raw.head(20), hide_index=True)

# ── Clean data ────────────────────────────────────────────
df = df_raw.copy()
for col in df.columns:
    if df[col].dtype == "object":
        df[col] = df[col].astype(str).str.strip()
empty_cells = int(df.isnull().sum().sum())
numeric_cols = df.select_dtypes(include="number").columns
df[numeric_cols] = df[numeric_cols].fillna(0)
df = df.fillna("N/A")

# ── ALL sidebar widgets OUTSIDE button ───────────────────
sort_col = st.sidebar.selectbox("Sort by", ["None"] + list(df.columns), key="sort_col")
sort_order = st.sidebar.selectbox("Sort order", ["Descending", "Ascending"], key="sort_order") if sort_col != "None" else "Descending"

num_cols = list(df.select_dtypes(include="number").columns)
cat_cols = list(df.select_dtypes(exclude="number").columns)

group_col = None
value_col = None
summary_sort = "None"
summary_order = "Descending"
chart_metric = "Total"

if mode == "Full Business Report" and num_cols and cat_cols:
    group_col = st.sidebar.selectbox("Group by", cat_cols, key="group_col")
    value_col = st.sidebar.selectbox("Analyze", num_cols, key="value_col")
    summary_sort = st.sidebar.selectbox("Sort summary by", ["None", "Total", "Average", "Highest", "Count"], key="summary_sort")
    if summary_sort != "None":
        summary_order = st.sidebar.selectbox("Summary order", ["Descending", "Ascending"], key="summary_order")
    chart_metric = st.sidebar.selectbox("Chart metric", ["Total", "Average", "Highest", "Count"], key="chart_metric")

# ── Generate Report Button ────────────────────────────────
if st.button("🚀 Generate Report"):

    prog = st.progress(0)

    with st.spinner("Processing..."):

        if sort_col != "None":
            try:
                df = df.sort_values(sort_col, ascending=(sort_order == "Ascending"))
            except Exception:
                st.warning("⚠️ Sorting failed.")

        prog.progress(40)

        st.subheader("📊 Clean Data")
        st.dataframe(df, hide_index=True)

        summary_df = None

        if mode == "Full Business Report" and group_col and value_col:
            df[value_col] = pd.to_numeric(df[value_col], errors="coerce")
            summary_df = df.groupby(group_col)[value_col].agg(
                Total="sum", Average="mean", Highest="max", Count="count"
            ).reset_index()
            summary_df["Total"] = summary_df["Total"].round(2)
            summary_df["Average"] = summary_df["Average"].round(2)
            summary_df["Highest"] = summary_df["Highest"].round(2)

            if summary_sort != "None":
                summary_df = summary_df.sort_values(summary_sort, ascending=(summary_order == "Ascending"))

            st.subheader("📈 Summary Report")
            st.dataframe(summary_df, hide_index=True)

            st.subheader(f"{chart_metric} by {group_col}")
            fig = px.bar(
                summary_df,
                x=group_col,
                y=chart_metric,
                labels={group_col: group_col, chart_metric: chart_metric},
                color_discrete_sequence=["#4A90D9"]
            )
            fig.update_layout(
                xaxis_title=group_col,
                yaxis_title=chart_metric,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(size=12),
                margin=dict(t=30, b=120)
            )
            fig.update_xaxes(tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)

        elif mode == "Full Business Report":
            st.warning("⚠️ Need at least one text and one numeric column for a summary.")

        prog.progress(80)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer) as writer:
            df.to_excel(writer, index=False, sheet_name="Scraped Data")
            if summary_df is not None:
                summary_df.to_excel(writer, index=False, sheet_name="Summary")

        buffer.seek(0)
        wb = load_workbook(buffer)
        for sheet in wb.sheetnames:
            for cell in wb[sheet][1]:
                cell.font = Font(bold=True)
        final = io.BytesIO()
        wb.save(final)
        final.seek(0)
        prog.progress(100)

    col1, col2, col3 = st.columns(3)
    col1.metric("Rows Extracted", len(df))
    col2.metric("Pages Scraped", total_pages)
    col3.metric("Empty Cells Fixed", empty_cells)
    st.success("✅ Report ready!")
    st.download_button("⬇️ Download Excel", final.getvalue(), "scraped_data.xlsx")
