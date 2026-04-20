from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd

# Read the CSV file
df = pd.read_csv("sales_q1.csv")

# Clean the data
df = df.drop_duplicates()
for col in df.columns:
    try:
        df[col] = df[col].str.strip().str.title()
    except:
        pass

# Drop rows with no Amount
df = df.dropna(subset=["Amount"])

# Calculate summary per salesperson
summary = df.groupby("Salesperson")["Amount"].agg(["sum", "mean", "max"])
summary.columns = ["Total", "Average", "Highest"]
summary["Average"] = summary["Average"].round()
# Find top salesperson
top = summary["Total"].idxmax()
summary=summary.reset_index()
print(f"Top salesperson: {top}")
print(summary)

# Save to Excel
summary.to_excel("pandas_report.xlsx", index=False)
print("Done!")

# Bold the header row
wb=load_workbook("pandas_report.xlsx")
ws=wb.active
for cell in ws[1]:
    cell.font=Font(bold=True)
wb.save("pandas_report.xlsx")