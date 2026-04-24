import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill
import time

# Scrape all 50 pages and collect book titles and prices
data = []
for page in range(1, 51):
    url = f"https://books.toscrape.com/catalogue/page-{page}.html"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    prices = soup.find_all("p", class_="price_color")
    titles = soup.find_all("h3")
    for price, title in zip(prices, titles):
        data.append({"Title": title.find("a")["title"], "Price": price.text})
    print(f"Scraped page {page}")
    time.sleep(0.5)

# Convert to DataFrame and clean the price column
df = pd.DataFrame(data)
df["Price"] = df["Price"].str.replace("Â£", "").str.replace("£", "").astype(float)

# Calculate summary statistics
cheapest = df.loc[df["Price"].idxmin()]
most_expensive = df.loc[df["Price"].idxmax()]
average = round(df["Price"].mean(), 2)

# Sort books by price and save to Excel
df = df.sort_values("Price")
df["Price"] = df["Price"].apply(lambda x: f"£{x:.2f}")
df.to_excel("price_report.xlsx", index=False)

# Style the main sheet — bold headers
wb = load_workbook("price_report.xlsx")
ws = wb.active
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="1D9E75", end_color="1D9E75", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    
# Auto width
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4
    
# Add summary sheet with key statistics
ws2 = wb.create_sheet("Summary")
ws2.append(["Metric", "Value"])
ws2.append(["Total Books", len(df)])
ws2.append(["Average Price", f"£{average}"])
ws2.append(["Cheapest Book", f"{cheapest['Title']} — £{cheapest['Price']}"])
ws2.append(["Most Expensive", f"{most_expensive['Title']} — £{most_expensive['Price']}"])
for cell in ws2[1]:
    cell.font = Font(bold=True)
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 60

wb.save("price_report.xlsx")
print(f"Done! Scraped {len(df)} books.")