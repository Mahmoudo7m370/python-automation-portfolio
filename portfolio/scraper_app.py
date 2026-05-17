import io
import re
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import requests
from bs4 import BeautifulSoup

# ── Page Header ───────────────────────────────────────────────
st.title("Web Data Extractor")
st.write("Extract structured data from any website and download a clean Excel report.")

# ── Sidebar ───────────────────────────────────────────────────
st.sidebar.header("Settings")

mode = st.sidebar.selectbox(
    "Report Mode",
    ["Extract Data Only", "Full Business Report"]
)

scraping_method = st.sidebar.radio(
    "Scraping Mode",
    [
        "Auto (Tables)",
        "Smart Extract (Books / Products)",
        "Table Parser (Wikipedia / Any Table)",
        "Advanced (CSS Selector)"
    ]
)

target_url = st.sidebar.text_input("Enter Website URL")

if not target_url:
    st.info("Enter a URL to begin.")
    st.stop()

# ── Fetch Page ────────────────────────────────────────────────
try:
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    response = requests.get(target_url, timeout=15, headers=headers)
    html = response.text
except Exception:
    st.error("❌ Could not access this URL. Check the link and try again.")
    st.stop()

if response.status_code != 200:
    st.error(f"❌ Access denied (Status {response.status_code})")
    st.stop()

st.success(f"✅ Page loaded successfully ({len(html):,} characters)")

scraped_df = None

# ══════════════════════════════════════════════════════════════
# MODE 1 — Auto Tables (pd.read_html)
# ══════════════════════════════════════════════════════════════
if scraping_method == "Auto (Tables)":
    try:
        tables = pd.read_html(io.StringIO(html))
        st.success(f"✅ Found {len(tables)} table(s)")
        table_index = st.sidebar.selectbox(
            "Select Table",
            range(len(tables)),
            format_func=lambda x: f"Table {x+1} ({len(tables[x])} rows)"
        )
        scraped_df = tables[table_index]
    except Exception:
        st.warning("⚠️ No tables detected. Try Table Parser mode.")

# ══════════════════════════════════════════════════════════════
# MODE 2 — Smart Extract (Books / Products)
# ══════════════════════════════════════════════════════════════
elif scraping_method == "Smart Extract (Books / Products)":
    soup = BeautifulSoup(html, "html.parser")
    products = soup.select("article.product_pod")

    if products:
        rating_map = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}
        rows = []

        for p in products:
            title_tag = p.select_one("h3 a")
            title = title_tag["title"] if title_tag and title_tag.has_attr("title") else (title_tag.get_text(strip=True) if title_tag else "N/A")

            price_tag = p.select_one("p.price_color")
            price_text = price_tag.get_text(strip=True) if price_tag else "0"
            price = float(re.sub(r"[^\d.]", "", price_text)) if price_text else 0.0

            rating_tag = p.select_one("p.star-rating")
            rating_word = rating_tag["class"][1] if rating_tag and len(rating_tag["class"]) > 1 else "Zero"
            rating = rating_map.get(rating_word, 0)

            avail_tag = p.select_one("p.availability")
            availability = avail_tag.get_text(strip=True) if avail_tag else "N/A"

            rows.append({
                "Title": title,
                "Price (£)": price,
                "Rating": rating,
                "Availability": availability
            })

        scraped_df = pd.DataFrame(rows)
        st.success(f"✅ Extracted {len(scraped_df)} products")
    else:
        st.error("❌ No products found. This mode works best with books.toscrape.com.")

# ══════════════════════════════════════════════════════════════
# MODE 3 — Table Parser (Wikipedia / Any HTML Table)
# Finds tables using BeautifulSoup, parses tr/td into columns.
# tr = table row, td/th = table cell.
# Same concept as iter_rows() in openpyxl — just on HTML.
# ══════════════════════════════════════════════════════════════
elif scraping_method == "Table Parser (Wikipedia / Any Table)":

    soup = BeautifulSoup(html, "html.parser")
    all_tables = soup.find_all("table")

    if not all_tables:
        st.error("❌ No tables found on this page.")
        st.stop()

    st.info(f"Found {len(all_tables)} table(s) on this page.")

    # Build preview labels for each table
    table_labels = []
    for i, t in enumerate(all_tables):
        first_row = t.find("tr")
        if first_row:
            cells = [c.get_text(strip=True)[:20] for c in first_row.find_all(["th", "td"])[:4]]
            label = f"Table {i+1}: {', '.join(cells)}"
        else:
            label = f"Table {i+1}"
        table_labels.append(label)

    selected_index = st.sidebar.selectbox(
        "Select Table",
        range(len(all_tables)),
        format_func=lambda x: table_labels[x]
    )
    selected_table = all_tables[selected_index]

    # Parse tr rows → DataFrame
    rows = []
    headers = []
    all_rows = selected_table.find_all("tr")

    for row in all_rows:
        cells = row.find_all(["th", "td"])
        cell_values = [c.get_text(strip=True) for c in cells]

        if not cell_values:
            continue

        if not headers:
            headers = cell_values
        else:
            while len(cell_values) < len(headers):
                cell_values.append("")
            rows.append(cell_values[:len(headers)])

    if not rows:
        st.error("❌ Could not parse any rows from this table.")
        st.stop()

    # Clean duplicate or empty header names
    clean_headers = []
    seen = {}
    for h in headers:
        h = h.strip() if h.strip() else "Column"
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0
        clean_headers.append(h)

    scraped_df = pd.DataFrame(rows, columns=clean_headers)

    # Auto-convert numeric columns — strip footnotes like [1], commas, symbols
    for col in scraped_df.columns:
        try:
            cleaned = scraped_df[col].str.replace(r"\[.*?\]", "", regex=True)
            cleaned = cleaned.str.replace(r"[,$€£%,†♠]", "", regex=True).str.strip()
            scraped_df[col] = pd.to_numeric(cleaned, errors="raise")
        except Exception:
            pass

    st.success(f"✅ Parsed {len(scraped_df)} rows and {len(scraped_df.columns)} columns")

# ══════════════════════════════════════════════════════════════
# MODE 4 — Advanced CSS Selector
# ══════════════════════════════════════════════════════════════
elif scraping_method == "Advanced (CSS Selector)":
    selector = st.text_input("Enter CSS Selector (e.g. h3.product-title, div.price)")

    if not selector:
        st.info("Enter a CSS selector to continue.")
        st.stop()

    soup = BeautifulSoup(html, "html.parser")
    elements = soup.select(selector)

    if not elements:
        st.error("❌ No elements found with that selector.")
        st.stop()

    data = [el.get_text(strip=True) for el in elements]
    scraped_df = pd.DataFrame(data, columns=["Extracted Data"])
    st.success(f"✅ Extracted {len(scraped_df)} items")

# ══════════════════════════════════════════════════════════════
# PREVIEW + PROCESS
# ══════════════════════════════════════════════════════════════
if scraped_df is not None and len(scraped_df) > 0:

    st.subheader("🔍 Raw Preview")
    st.dataframe(scraped_df.head(20), hide_index=True)

    if st.button("🚀 Generate Report"):

        progress = st.progress(0)

        with st.spinner("Processing..."):

            df = scraped_df.copy()
            progress.progress(20)

            for col in df.columns:
                if df[col].dtype == "object":
                    df[col] = df[col].astype(str).str.strip()

            empty_cells = int(df.isnull().sum().sum())

            numeric_cols = df.select_dtypes(include="number").columns
            df[numeric_cols] = df[numeric_cols].fillna(0)
            df = df.fillna("N/A")

            progress.progress(40)

            sort_col = st.sidebar.selectbox("Sort by", ["None"] + list(df.columns), key="sort_col")
            if sort_col != "None":
                order = st.sidebar.selectbox("Order", ["Descending", "Ascending"], key="sort_order")
                try:
                    df = df.sort_values(sort_col, ascending=(order == "Ascending"))
                except Exception:
                    st.warning("⚠️ Could not sort — mixed data types.")

            progress.progress(60)

            st.subheader("📊 Clean Data")
            st.dataframe(df, hide_index=True)

            summary_df = None

            if mode == "Full Business Report":
                num_cols = list(df.select_dtypes(include="number").columns)
                cat_cols = list(df.select_dtypes(exclude="number").columns)

                if not num_cols or not cat_cols:
                    st.warning("⚠️ Need at least one text column and one numeric column for a summary.")
                else:
                    group_col = st.sidebar.selectbox("Group by", cat_cols, key="group_col")
                    value_col = st.sidebar.selectbox("Analyze", num_cols, key="value_col")

                    df[value_col] = pd.to_numeric(df[value_col], errors="coerce")

                    summary_df = df.groupby(group_col)[value_col].agg(
                        Total="sum",
                        Average="mean",
                        Highest="max",
                        Count="count"
                    ).reset_index()

                    summary_df["Total"] = summary_df["Total"].round(2)
                    summary_df["Average"] = summary_df["Average"].round(2)
                    summary_df["Highest"] = summary_df["Highest"].round(2)

                    summary_sort = st.sidebar.selectbox(
                        "Sort summary by",
                        ["None", "Total", "Average", "Highest", "Count"],
                        key="summary_sort"
                    )
                    if summary_sort != "None":
                        summary_order = st.sidebar.selectbox(
                            "Summary order",
                            ["Descending", "Ascending"],
                            key="summary_order"
                        )
                        summary_df = summary_df.sort_values(
                            summary_sort,
                            ascending=(summary_order == "Ascending")
                        )

                    st.subheader("📈 Summary Report")
                    st.dataframe(summary_df, hide_index=True)

                    chart_metric = st.selectbox(
                        "Chart metric",
                        ["Total", "Average", "Highest", "Count"],
                        key="chart_metric"
                    )
                    chart_data = summary_df.set_index(group_col)[[chart_metric]]
                    st.subheader(f"{chart_metric} by {group_col}")
                    st.bar_chart(chart_data)

            progress.progress(80)

            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer) as writer:
                df.to_excel(writer, index=False, sheet_name="Scraped Data")
                if summary_df is not None:
                    summary_df.to_excel(writer, index=False, sheet_name="Summary")

            buffer.seek(0)
            wb = load_workbook(buffer)

            for sheet in wb.sheetnames:
                for cell in wb[sheet][1]:
                    cell.font = Font(bold=True)

            final = io.BytesIO()
            wb.save(final)
            final.seek(0)
            progress.progress(100)

        col1, col2 = st.columns(2)
        col1.metric("Rows Extracted", len(df))
        col2.metric("Empty Cells Fixed", empty_cells)

        st.success("✅ Your report is ready!")
        st.download_button("⬇️ Download Excel", final.getvalue(), "scraped_data.xlsx")

else:
    st.info("No data extracted yet.")
