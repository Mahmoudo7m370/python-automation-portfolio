from openpyxl import load_workbook
wb= load_workbook("clients.xlsx")
ws=wb["Clients"]
count=0
for row in ws.iter_rows(min_row=2,max_row=6,values_only=True):
    if (row[3])>100:
        print(row)
        count+=1
print(count)
       
