from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill
import pandas as pd
import glob
# Read all of the CSV files
all_files=glob.glob("*.csv")
if not all_files:
    print("No CSV files found!")
    exit()
df=pd.concat((pd.read_csv(f) for f in all_files),ignore_index=True)

# Clean the data
df = df.drop_duplicates()
for col in df.columns:
    try:
        df[col] = df[col].str.strip().str.title()
    except:
        pass

# Ensure Amount is numeric
df["Amount"]=pd.to_numeric(df["Amount"],errors="coerce")

# Drop rows with no Amount
df = df.dropna(subset=["Amount"])

# Calculate summary per salesperson
summary = df.groupby("Salesperson")["Amount"].agg(Total="sum", Average="mean", Highest="max").reset_index()
summary["Average"] = summary["Average"].round()

#Sort data of total
summary=summary.sort_values("Total",ascending=False)

# Find top salesperson
top = summary.iloc[0]["Salesperson"]
print(f"Top salesperson: {top}")
print(summary)

# Save to Excel
summary.to_excel("pandas_report.xlsx", index=False)
print("Done!")

# Bold the header row
wb=load_workbook("pandas_report.xlsx")
ws=wb.active
for cell in ws[1]:
    cell.font=Font(bold=True)

#Hilighting the top person with green 
green=PatternFill(start_color="90EE90",end_color="90EE90",fill_type="solid")
for cell in ws[2]:
    cell.fill=green
wb.save("pandas_report.xlsx")