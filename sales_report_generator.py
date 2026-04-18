import csv
import glob
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
files=glob.glob("*.csv")
files = [f for f in files if f != "final_report.xlsx"]
if len(files) == 0:
    print("Error:no file file with csv type ")
    exit()
rows=[]
header=None
for file in files:
    with open(f"{file}","r") as f:
        reader = csv.reader(f)
        header_row=next(reader)
        if "Amount" not in header_row:
            continue
        if "Salesperson"not in header_row:
            continue
        if header is None:
            header=header_row
            i=header.index("Amount")
            i2=header.index("Salesperson")
        for row in reader:
            rows.append(row)
if header is None:
    print("Error: No files with valid headers")
    exit()
unique=[]
cleaned_rows=[]
for row in rows:
    cleaned=[]
    for cell in row:
        if cell =="":
            cell="N/A"
        elif  isinstance(cell,str):
            cell=cell.strip().title()
        cleaned.append(cell)
    cleaned_rows.append(tuple(cleaned))
seen=set()
summary={}
for row in cleaned_rows:
    if row not in seen:
        seen.add(row)
        unique.append(row)
for row in unique:
    name=row[i2]
    amount=row[i]
    if amount=="N/A":
        continue
    if name not in summary:
        summary[name]=[]
    summary[name].append(int(amount))
new_header=("Salesperson","Total","Average","Highest")
wb=Workbook()
ws=wb.active
ws.append(new_header)
for name,amounts in summary.items():
    Total = sum(amounts)
    Average=round(sum(amounts)/len(amounts))
    Highest=max(amounts)
    new_row=(name,Total,Average,Highest)
    ws.append(new_row)
for cell in ws[1]:
    cell.font=Font(bold=True)
green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
top_salesperson = max(summary, key=lambda name: sum(summary[name]))
for row in ws.iter_rows(min_row=2):
    if row[0].value == top_salesperson:
        for cell in row:
            cell.fill = green
ws.freeze_panes="A2"
wb.save("final_report.xlsx")


        
        
